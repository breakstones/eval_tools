"""API routes for evaluation task management."""

import asyncio
import json
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Query, WebSocket, WebSocketDisconnect, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_eval_service
from app.api.websocket import manager as ws_manager
from app.database import get_db
from app.schemas.eval import (
    EvalResultResponse,
    EvalRunResponse,
    EvalTaskCreate,
    EvalTaskResponse,
    EvalTaskUpdate,
    TemplateTestRequest,
    TemplateTestResponse,
)
from app.services.eval_service import EvalService

router = APIRouter(prefix="/api/eval", tags=["evaluation"])


@router.get("/tasks", response_model=List[EvalTaskResponse])
async def get_eval_tasks(
    set_id: Optional[str] = Query(None),
    service: EvalService = Depends(get_eval_service),
) -> List[EvalTaskResponse]:
    """Get all evaluation tasks."""
    tasks = await service.get_eval_tasks(set_id)
    result = []
    for t in tasks:
        # Get model details
        model_details = await service.get_model_with_provider(t.model_id)
        model_config = None
        if model_details:
            model_config = {
                "model_code": model_details["model_code"],
                "display_name": model_details["display_name"],
                "provider": {
                    "id": model_details["provider"]["id"],
                    "name": model_details["provider"]["name"],
                },
            }
        result.append(EvalTaskResponse(
            id=t.id,
            name=t.name,
            set_id=t.set_id,
            model_id=t.model_id,
            concurrency=getattr(t, 'concurrency', 1),
            request_template=t.request_template_dict,
            system_prompt=t.system_prompt,
            model_info=model_config,
            status=t.status,
            summary=t.summary_dict,
            created_at=t.created_at,
            updated_at=t.updated_at,
        ))
    return result


@router.get("/tasks/{task_id}", response_model=EvalTaskResponse)
async def get_eval_task(
    task_id: str,
    service: EvalService = Depends(get_eval_service),
) -> EvalTaskResponse:
    """Get an evaluation task by ID."""
    task = await service.get_eval_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="评测任务不存在")

    # Get model details
    model_details = await service.get_model_with_provider(task.model_id)
    model_config = None
    if model_details:
        model_config = {
            "model_code": model_details["model_code"],
            "display_name": model_details["display_name"],
            "provider": {
                "id": model_details["provider"]["id"],
                "name": model_details["provider"]["name"],
            },
        }

    return EvalTaskResponse(
        id=task.id,
        name=task.name,
        set_id=task.set_id,
        model_id=task.model_id,
        concurrency=getattr(task, 'concurrency', 1),
        request_template=task.request_template_dict,
        system_prompt=task.system_prompt,
        model_info=model_config,
        status=task.status,
        summary=task.summary_dict,
        created_at=task.created_at,
        updated_at=task.updated_at,
    )


@router.post("/tasks", response_model=EvalTaskResponse, status_code=201)
async def create_eval_task(
    data: EvalTaskCreate,
    service: EvalService = Depends(get_eval_service),
) -> EvalTaskResponse:
    """Create a new evaluation task."""
    try:
        task = await service.create_eval_task(data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Get model details
    model_details = await service.get_model_with_provider(task.model_id)
    model_config = None
    if model_details:
        model_config = {
            "model_code": model_details["model_code"],
            "display_name": model_details["display_name"],
            "provider": {
                "id": model_details["provider"]["id"],
                "name": model_details["provider"]["name"],
            },
        }

    return EvalTaskResponse(
        id=task.id,
        name=task.name,
        set_id=task.set_id,
        model_id=task.model_id,
        concurrency=getattr(task, 'concurrency', 1),
        request_template=task.request_template_dict,
        system_prompt=task.system_prompt,
        model_info=model_config,
        status=task.status,
        summary=None,
        created_at=task.created_at,
        updated_at=task.updated_at,
    )


@router.delete("/tasks/{task_id}", status_code=204)
async def delete_eval_task(
    task_id: str,
    service: EvalService = Depends(get_eval_service),
) -> None:
    """Delete an evaluation task."""
    deleted = await service.delete_eval_task(task_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="评测任务不存在")


@router.put("/tasks/{task_id}", response_model=EvalTaskResponse)
async def update_eval_task(
    task_id: str,
    data: EvalTaskUpdate,
    service: EvalService = Depends(get_eval_service),
) -> EvalTaskResponse:
    """Update an evaluation task."""
    # Debug logging
    print(f"[DEBUG] Update task {task_id} received:")
    print(f"[DEBUG]   Raw data dict: {data.model_dump()}")
    print(f"[DEBUG]   name={data.name}")
    print(f"[DEBUG]   model_id={data.model_id}")
    print(f"[DEBUG]   concurrency={data.concurrency} (type: {type(data.concurrency).__name__ if data.concurrency is not None else 'NoneType'})")
    print(f"[DEBUG]   request_template={data.request_template}")
    print(f"[DEBUG]   system_prompt length={len(data.system_prompt) if data.system_prompt else 0}")

    try:
        request_template = data.request_template.model_dump() if data.request_template else None
        print(f"[DEBUG] Parsed request_template: {request_template}")
        task = await service.update_eval_task(
            task_id,
            data.model_id,
            request_template,
            data.system_prompt,
            data.concurrency,
            data.name
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Get model details
    model_details = await service.get_model_with_provider(task.model_id)
    model_config = None
    if model_details:
        model_config = {
            "model_code": model_details["model_code"],
            "display_name": model_details["display_name"],
            "provider": {
                "id": model_details["provider"]["id"],
                "name": model_details["provider"]["name"],
            },
        }

    return EvalTaskResponse(
        id=task.id,
        name=task.name,
        set_id=task.set_id,
        model_id=task.model_id,
        concurrency=getattr(task, 'concurrency', 1),
        request_template=task.request_template_dict,
        system_prompt=task.system_prompt,
        model_info=model_config,
        status=task.status,
        summary=task.summary_dict,
        created_at=task.created_at,
        updated_at=task.updated_at,
    )


@router.post("/tasks/{task_id}/test-template", response_model=TemplateTestResponse)
async def test_template(
    task_id: str,
    data: TemplateTestRequest,
    service: EvalService = Depends(get_eval_service),
) -> TemplateTestResponse:
    """Test request template rendering and optionally send a real request."""
    try:
        rendered_request, actual_response, error = await service.test_template(
            task_id,
            data.case_id,
            data.test_input,
        )
        return TemplateTestResponse(
            rendered_request=rendered_request,
            actual_response=actual_response,
            error=error,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/tasks/{task_id}/results", response_model=List[EvalResultResponse])
async def get_eval_results(
    task_id: str,
    service: EvalService = Depends(get_eval_service),
) -> List[EvalResultResponse]:
    """Get evaluation results for a task (latest run results)."""
    results = await service.get_eval_results(task_id)
    return [
        EvalResultResponse(
            id=result.id,
            run_id=result.run_id,
            task_id=result.task_id,
            case_id=result.case_id,
            case_uid=case.case_uid,
            actual_output=result.actual_output,
            is_passed=result.is_passed,
            execution_error=getattr(result, 'execution_error', None),
            evaluator_logs=result.evaluator_logs_list,
            execution_duration=getattr(result, 'execution_duration', None),
            skill_tokens=getattr(result, 'skill_tokens', None),
            evaluator_tokens=getattr(result, 'evaluator_tokens', None),
            created_at=result.created_at,
        )
        for result, case in results
    ]


@router.get("/results/{result_id}", response_model=EvalResultResponse)
async def get_eval_result(
    result_id: str,
    db: AsyncSession = Depends(get_db),
    service: EvalService = Depends(get_eval_service),
) -> EvalResultResponse:
    """Get a single evaluation result."""
    result_data = await service.get_eval_result(result_id)
    if result_data is None:
        raise HTTPException(status_code=404, detail="评测结果不存在")

    result, case = result_data
    return EvalResultResponse(
        id=result.id,
        run_id=result.run_id,
        task_id=result.task_id,
        case_id=result.case_id,
        case_uid=case.case_uid,
        actual_output=result.actual_output,
        is_passed=result.is_passed,
        execution_error=getattr(result, 'execution_error', None),
        evaluator_logs=result.evaluator_logs_list,
        execution_duration=getattr(result, 'execution_duration', None),
        skill_tokens=getattr(result, 'skill_tokens', None),
        evaluator_tokens=getattr(result, 'evaluator_tokens', None),
        created_at=result.created_at,
    )


@router.get("/tasks/{task_id}/runs", response_model=List[EvalRunResponse])
async def get_eval_runs(
    task_id: str,
    service: EvalService = Depends(get_eval_service),
) -> List[EvalRunResponse]:
    """Get all runs for an evaluation task."""
    # Verify task exists
    task = await service.get_eval_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="评测任务不存在")

    runs = await service.get_eval_runs(task_id)
    return [
        EvalRunResponse(
            id=r.id,
            task_id=r.task_id,
            run_number=r.run_number,
            status=r.status,
            summary=json.loads(r.summary) if r.summary else None,
            started_at=r.started_at,
            completed_at=r.completed_at,
            error=r.error,
            total_duration_ms=getattr(r, 'total_duration_ms', None),
            total_skill_tokens=getattr(r, 'total_skill_tokens', None),
            total_evaluator_tokens=getattr(r, 'total_evaluator_tokens', None),
        )
        for r in runs
    ]


@router.get("/runs/{run_id}", response_model=EvalRunResponse)
async def get_eval_run(
    run_id: str,
    service: EvalService = Depends(get_eval_service),
) -> EvalRunResponse:
    """Get a single evaluation run by ID."""
    run = await service.get_eval_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="运行记录不存在")

    return EvalRunResponse(
        id=run.id,
        task_id=run.task_id,
        run_number=run.run_number,
        status=run.status,
        summary=json.loads(run.summary) if run.summary else None,
        started_at=run.started_at,
        completed_at=run.completed_at,
        error=run.error,
        total_duration_ms=getattr(run, 'total_duration_ms', None),
        total_skill_tokens=getattr(run, 'total_skill_tokens', None),
        total_evaluator_tokens=getattr(run, 'total_evaluator_tokens', None),
    )


@router.post("/tasks/{task_id}/rerun", response_model=EvalRunResponse, status_code=201)
async def rerun_eval_task(
    task_id: str,
    background_tasks: BackgroundTasks,
    service: EvalService = Depends(get_eval_service),
) -> EvalRunResponse:
    """Start a new evaluation run (runs asynchronously)."""
    # Create a new run
    run = await service.create_eval_run(task_id)

    # Start evaluation in background task
    background_tasks.add_task(run_evaluation_async, service, run.id, task_id)

    # Return the run info immediately
    return EvalRunResponse(
        id=run.id,
        task_id=run.task_id,
        run_number=run.run_number,
        status=run.status,
        summary=None,
        started_at=run.started_at,
        completed_at=None,
        error=None,
    )


async def run_evaluation_async(
    service: EvalService,
    run_id: str,
    task_id: str,
) -> None:
    """Run evaluation in background and broadcast updates via WebSocket.

    Args:
        service: The EvalService instance
        run_id: The run ID
        task_id: The task ID
    """
    import traceback
    print(f"[DEBUG] run_evaluation_async started: run_id={run_id}, task_id={task_id}")
    try:
        # Broadcast run created event
        print(f"[DEBUG] Broadcasting run_created event")
        await ws_manager.broadcast_event(
            task_id,
            "run_created",
            {
                "run_id": run_id,
                "run_number": await service.get_run_number(run_id),
            },
        )

        # Run the evaluation
        print(f"[DEBUG] Calling run_evaluation_with_ws")
        summary = await service.run_evaluation_with_ws(
            run_id,
            task_id,
            ws_manager,
        )
        print(f"[DEBUG] Evaluation completed: {summary}")

        # Broadcast completion event
        await ws_manager.broadcast_event(
            task_id,
            "complete",
            {"summary": summary},
        )

    except Exception as e:
        print(f"[ERROR] Evaluation failed: {e}")
        traceback.print_exc()
        # Broadcast error event
        try:
            await ws_manager.broadcast_event(
                task_id,
                "error",
                {"error": str(e)},
            )
        except Exception as ws_error:
            print(f"[ERROR] Failed to broadcast error: {ws_error}")


@router.websocket("/ws/eval/{task_id}")
async def websocket_eval_updates(
    task_id: str,
    websocket: WebSocket,
) -> None:
    """WebSocket endpoint for real-time evaluation updates.

    Args:
        task_id: The task ID to listen for updates
        websocket: The WebSocket connection
    """
    await ws_manager.connect(websocket, task_id)

    try:
        # Send initial connection confirmation
        await websocket.send_json({
            "type": "connected",
            "data": {"task_id": task_id},
        })

        # Keep connection alive and handle incoming messages
        while True:
            # Receive any messages from client (for ping/pong, etc.)
            data = await websocket.receive_text()

            # Handle ping/pong to keep connection alive
            if data == "ping":
                await websocket.send_text("pong")

    except WebSocketDisconnect:
        ws_manager.disconnect(websocket, task_id)
    except Exception as e:
        ws_manager.disconnect(websocket, task_id)
        raise


@router.get("/stream/{task_id}")
async def stream_evaluation(
    task_id: str,
    service: EvalService = Depends(get_eval_service),
) -> StreamingResponse:
    """Stream evaluation progress via Server-Sent Events (deprecated - use WebSocket instead)."""

    async def event_stream():
        async for event in service.stream_evaluation(task_id):
            yield event

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/runs/{run_id}/results", response_model=List[EvalResultResponse])
async def get_run_results(
    run_id: str,
    service: EvalService = Depends(get_eval_service),
) -> List[EvalResultResponse]:
    """Get evaluation results for a specific run."""
    results = await service.get_run_results(run_id)
    return [
        EvalResultResponse(
            id=result.id,
            run_id=result.run_id,
            task_id=result.task_id,
            case_id=result.case_id,
            case_uid=case.case_uid,
            actual_output=result.actual_output,
            is_passed=result.is_passed,
            execution_error=getattr(result, 'execution_error', None),
            evaluator_logs=result.evaluator_logs_list,
            execution_duration=getattr(result, 'execution_duration', None),
            skill_tokens=getattr(result, 'skill_tokens', None),
            evaluator_tokens=getattr(result, 'evaluator_tokens', None),
            created_at=result.created_at,
        )
        for result, case in results
    ]


@router.get("/runs/{run_id}/export")
async def export_run_results(
    run_id: str,
    failed_only: bool = Query(False, description="仅导出不通过的结果"),
    service: EvalService = Depends(get_eval_service),
):
    """导出评测结果为 Excel 文件."""
    import pandas as pd
    import openpyxl
    from io import BytesIO
    from datetime import datetime

    # 获取评测结果
    results = await service.get_run_results(run_id)

    # 如果需要仅不通过的，进行过滤
    if failed_only:
        results = [(r, c) for r, c in results if not r.is_passed or r.execution_error]

    # 先收集所有评估器名称，确定列结构
    all_evaluator_names = set()
    for result, case in results:
        evaluator_logs = result.evaluator_logs_list
        for log in evaluator_logs:
            all_evaluator_names.add(log.get("evaluator", ""))

    # 排序评估器名称以确保列顺序一致
    evaluator_names = sorted([name for name in all_evaluator_names if name])
    max_evaluators = len(evaluator_names)

    # 如果没有评估器，使用默认列
    if max_evaluators == 0:
        max_evaluators = 1

    # 构建 Excel 数据 - 每个用例占一行，评估器按列展示
    excel_data = []
    for result, case in results:
        row_data = {
            "用例编号": case.case_uid or "",
            "用例用户输入": case.user_input or "",
            "用例预期输出": case.expected_output or "",
            "实际输出": result.actual_output or "",
            "评测结果": "通过" if result.is_passed else "不通过",
            "评测耗时(ms)": result.execution_duration if result.execution_duration is not None else "",
            "技能tokens": result.skill_tokens if result.skill_tokens is not None else "",
            "评估器tokens": result.evaluator_tokens if result.evaluator_tokens is not None else "",
        }

        # 为每个评估器添加列
        evaluator_logs = result.evaluator_logs_list
        evaluator_dict = {log.get("evaluator", ""): log for log in evaluator_logs}

        for idx, eval_name in enumerate(evaluator_names):
            log = evaluator_dict.get(eval_name, {})
            row_data[f"评估器{idx+1}名称"] = eval_name
            row_data[f"评估器{idx+1}结果"] = "通过" if log.get("passed", False) else "不通过"
            row_data[f"评估器{idx+1}原因"] = log.get("reason", "")

        excel_data.append(row_data)

    # 创建 DataFrame
    df = pd.DataFrame(excel_data)

    # 调整列顺序：基础列在前，评估器列在后
    base_columns = ["用例编号", "用例用户输入", "用例预期输出", "实际输出", "评测结果"]
    metric_columns = ["评测耗时(ms)", "技能tokens", "评估器tokens"]
    evaluator_columns = []
    for idx in range(max_evaluators):
        evaluator_columns.extend([f"评估器{idx+1}名称", f"评估器{idx+1}结果", f"评估器{idx+1}原因"])

    column_order = base_columns + evaluator_columns + metric_columns
    df = df[column_order]

    # 创建 Excel 文件
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='评测结果')

        # 获取工作表并设置列宽
        worksheet = writer.sheets['评测结果']

        # 动态设置列宽
        column_widths = {}
        for idx, col in enumerate(df.columns, start=1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            if "编号" in col or "结果" in col and "耗时" not in col:
                column_widths[col_letter] = 15
            elif "输入" in col or "输出" in col or "原因" in col:
                column_widths[col_letter] = 30
            elif "耗时" in col or "tokens" in col:
                column_widths[col_letter] = 15
            else:
                column_widths[col_letter] = 20

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # 设置所有单元格样式：垂直居中、水平左对齐
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal='left',
                    vertical='center',
                    wrap_text=True
                )

        # 设置默认行高为36
        for row in worksheet.iter_rows(min_row=2):  # 从第2行开始（跳过标题行）
            worksheet.row_dimensions[row[0].row].height = 36

    output.seek(0)

    # 生成文件名（包含时间戳和run_id以确保唯一性）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"eval_results_{timestamp}_run{run_id[:8]}.xlsx"

    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


